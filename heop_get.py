# -*- coding: utf-8 -*-
import requests
from requests.auth import HTTPDigestAuth
from variable import *
import openpyxl
from public_function import getColValues
from lxml import etree
import threading
import sys
from openpyxl.styles import Font, colors, PatternFill, Border,Side
from datetime import datetime

reload(sys)
sys.setdefaultencoding('utf8')


class Device:
    """
    ip:设备ip
    username：设备登录账号
    password:设备密码
    nas_address:nas盘地址，选填
    nas_file_path:nas盘文件路径，选填
    restore_choose:是否简单恢复，选填
    record_channel:录像通道
    """

    # 初始化用户输入的ip、HTTP端口号、用户名、密码等
    def __init__(self, ip, username, password):
        self.port = '80'  # 设备HTTP端口号
        self.username = username  # 帐户名
        self.password = password  # 密码
        self.ip = ip  # nvr地址
        self.result = u'{0}：'.format(self.ip)

    # SD卡兼容性：检查设备是否上线
    def check_status(self):
        n = 0
        while True:
            try:
                n += 1
                put = requests.get('http://{0}:{1}{2}'.format(self.ip, self.port, check_status_url),
                                   auth=HTTPDigestAuth(self.username, self.password), timeout=5)
                put.raise_for_status()
                return True
            except Exception as e:
                # 若超过2分钟未检测到设备在线，报错
                if n >= 2:
                    print(u"{0}不在线，请检查！！".format(self.ip))
                    self.result += u'不在线，请检查'
                    return False

    def get_current_status(self):
        n = 0
        status = []
        while True:
            try:
                n += 1
                get = requests.get('http://{0}:{1}{2}'.format(self.ip, '80', '/ISAPI/Custom/OpenPlatform/App'),
                                   auth=HTTPDigestAuth(self.username, self.password), timeout=10)
                get.raise_for_status()
                root = etree.XML(get.text.encode('utf-8'))
                run_status = root.xpath("//*[local-name()='runStatus']")
                package_name = root.xpath("//*[local-name()='packageName']")
                for i in run_status:
                    status.append(i.text)
                if 'true' in status:
                    index = status.index('true')
                    result = package_name[index].text
                    if result == u'\u4eba\u8138\u6293\u62cd':
                        result = u'人脸抓拍'
                    if result == u'AI\u5f00\u653e\u5e73\u53f0':
                        result = u'AI开放平台'
                    self.result += result
                    return True
                else:
                    result = u'无'
                    self.result += result
                    return True
            except Exception as e:
                if n >= 2:
                    result = u'获取失败'
                    self.result += result
                    print(u"{0}获取APP状态失败！！".format(self.ip))
                    return False

    def test(self):
        if self.check_status():
            self.get_current_status()

# 读取Excel信息
class Excel:

    # 初始化
    def __init__(self):
        self.filename = 'devices/devices.xlsx'  # 读取的文件名
        self.ip = []                        # nvr地址
        self.username = []  # 帐户名
        self.password = []  # 密码

    def read(self):
        try:
        # 读取Excel
            data = openpyxl.load_workbook(self.filename)

            # 读取Sheet1表
            read = data[u'HEOP获取']
            # 获取设备IP与格式化次数
            self.ip = list(filter(None, getColValues(read, 1)))[1:]  # nvr地址
            self.username = list(filter(None, getColValues(read, 2)))[1:]  # nvr地址
            self.password =  list(filter(None, getColValues(read, 3)))[1:]  # nvr地址
            return True
        except Exception as e:
            print(e)
            return False

    # 执行结果写入
    def write(self, result):

        # 背景红色填充
        fill = PatternFill("solid", fgColor="ff0000")

        # 设置字体格式
        font = Font(name=u"微软雅黑", size=13, color=colors.BLACK)
        border = Border(left=Side(style='medium', color=colors.BLACK),
                        right=Side(style='medium', color=colors.BLACK),
                        top=Side(style='medium', color=colors.BLACK),
                        bottom=Side(style='medium', color=colors.BLACK))

        # 打开Excel，设置Sheet页名为Sheet
        data = openpyxl.Workbook()
        write = data['Sheet']

        # 遍历写入执行结果
        lenth = len(self.ip)
        for i in range(lenth):
            write.cell(row=i + 2, column=1).value = self.ip[i]
            write.cell(row=i + 2, column=2).value = result[i]
            if 'Failed' in result[i]:
                write.cell(row=i + 2, column=2).fill = fill

        # 为数据设置格式
        for cols in write['A1:B' + str(lenth + 1)]:
            for j in cols:
                j.font = font
                j.border = border

        font_deviceModel = Font(name=u"微软雅黑", size=14, color=colors.BLACK, bold=True)
        write["A1"].font = font_deviceModel
        write.column_dimensions['A'].width = 15
        write.cell(row=1, column=1).value = u"设备IP"

        write["B1"].font = font_deviceModel
        write.column_dimensions['B'].width = 18
        write.cell(row=1, column=2).value = u"测试结果"


        write.column_dimensions["A"].auto_size = True
        write.column_dimensions["B"].auto_size = True
        report_name = 'devices/heop_report_%s.xlsx' % datetime.now().strftime('%Y-%m-%d-%H-%M-%S')
        data.save(report_name)
        print(u'测试结果已导出至工具目录下的{0}'.format(report_name))


# 多线程执行
class MyThread(threading.Thread, Device):

    def __init__(self, ip, username, password):
        threading.Thread.__init__(self)
        Device.__init__(self, ip, username, password)

    def run(self):
        self.test()


# 从读取到生成报告总流程
def heop_get():
    try:
        print(u'开始执行')
        result_r = []
        excel = Excel()
        if excel.read():
            nums = len(excel.ip)
            li = []
            for i in range(nums):
                t = MyThread(excel.ip[i], excel.username[i], excel.password[i])
                li.append(t)
                t.start()
            for t in li:
                t.join()
            result = []
            for i in range(nums):
                result.append(li[i].result)
            excel.write(result)
            for i in li:
                result_r.append(i.result)
        return result_r
    except Exception as e:
        print(e)

if __name__ == '__main__':
    heop_get()
