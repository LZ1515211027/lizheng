# -*- coding: utf-8 -*-
import threading
import openpyxl
from openpyxl.styles import Font, colors, PatternFill, Border,Side
import requests                           # 请求相关的库
from requests.auth import HTTPDigestAuth  # 摘要认证
import time
from datetime import datetime
from public_function import getColValues
import paramiko
import re
import xmltodict


class Device(object):

    """
    ip:设备ip
    port:设备HTTP端口号
    username：设备登录账号
    password:设备密码
    cout:格式化次数
    """

    # 初始化用户输入的ip、HTTP端口号、用户名、密码等
    def __init__(self, ip, username, password, port):
        self.ip = ip                      # 设备IP地址
        self.port = int(port)                  # 设备HTTP端口号
        self.username = username          # 用户名
        self.password = password          # 密码
        self.result = ''                  # 执行结果
        self.result_failed = ''           # 失败原因
        self.CurZoomPos = ''              # 变倍数据
        self.CurFocusPos = ''             # 聚焦数据
        self.storage = []                 # 设备存储介质个数
        self.result_r = u'执行通过'        # 打印的执行结果
        self.model = ''
        # 实例化SSHClient
        self.client = paramiko.SSHClient()

        # 自动添加策略，保存服务器的主机名和密钥信息，如果不添加，那么不再本地know_hosts文件中记录的主机将无法连接
        self.client.set_missing_host_key_policy(paramiko.AutoAddPolicy())

    # 登录设备
    def check(self):
        n = 0
        while True:
            try:
                n += 1
                get = requests.get(u'http://{0}:{1}{2}'.format(self.ip, '80', u'/ISAPI/System/deviceInfo'),
                                   auth=HTTPDigestAuth(self.username, self.password), timeout=5)
                get.raise_for_status()
                return True
            except Exception as e:
                if n >= 2:
                    self.result += u'Failed'
                    self.result_r = u"{0}：登录失败，请检查是否在线".format(self.ip)
                    self.result_failed += u'设备登录失败，请检查'
                    return False


    # 登录设备
    def check_z(self):
        n = 0
        while True:
            try:
                n += 1
                get = requests.get(u'http://{0}:{1}{2}'.format(self.ip, '80', u'/ISAPI/System/deviceInfo'),
                                   auth=HTTPDigestAuth(self.username, self.password), timeout=5)
                get.raise_for_status()
                get_text = xmltodict.parse(get.text)
                self.model = get_text['DeviceInfo']['model']
                if 'Z' in self.model:
                    return True
                else:
                    self.result += u'Pass'
                    self.result_failed += u'非电动镜头设备'
                    return False
            except Exception as e:
                if n >= 2:
                    self.result += u'Failed'
                    self.result_r = u"{0}：无法查询到设备是否为电动镜头".format(self.ip)
                    self.result_failed += u'无法查询到设备是否为电动镜头'
                    return False

    def open_ssh(self):
        n = 0
        while True:
            try:
                n += 1
                get = requests.get(u'http://{0}:{1}{2}'.format(self.ip, '80', u'/ISAPI/System/Network/ssh'),
                                   auth=HTTPDigestAuth(self.username, self.password), timeout=5)
                get.raise_for_status()
                get_text = xmltodict.parse(get.text)
                if get_text['SSH']['enabled'] == 'true':
                    pass
                else:
                    get_text['SSH']['enabled'] = 'true'
                    put_text = xmltodict.unparse(get_text)
                    put = requests.put(u'http://{0}:{1}{2}'.format(self.ip, '80', u'/ISAPI/System/Network/ssh'),data=put_text,
                                       auth=HTTPDigestAuth(self.username, self.password), timeout=5)
                    put.raise_for_status()
                return True
            except Exception as e:
                if n >= 2:
                    self.result += u'Failed'
                    self.result_r = u"{0}：开启SSH失败".format(self.ip)
                    self.result_failed += u'开启SSH失败'
                    return False

    # 获取设备电动镜头信息
    def check_position(self):
        try:
            self.client.connect(hostname=self.ip, port=self.port, username=self.username, password=self.password)
            channel = self.client.invoke_shell()
            time.sleep(1)
            channel.send(u'outputClose'.encode("utf-8") + u'\n'.encode("utf-8"))
            time.sleep(1)
            channel.send(u'outputOpen'.encode("utf-8") + u'\n'.encode("utf-8"))
            time.sleep(1)
            channel.send(u'camCmd 1  810a0c06ff'.encode("utf-8") + u'\n'.encode("utf-8"))
            time.sleep(1)
            stdout = channel.recv(1024 * 1024)
            # 关闭连接
            self.client.close()
            a = re.search('CurZoomPos=[1-9]\d*,CurFocusPos=[1-9]\d*', stdout)  # None
            b = a.group().split(',')
            self.CurZoomPos = b[0][11:]
            self.CurFocusPos = b[1][12:]
            self.result += u'Pass'
            return True
        except Exception as e:
            self.result += u'Failed'
            self.result_r = u"{0}：SSH获取电动镜头信息失败".format(self.ip)
            self.result_failed += u'SSH获取电动镜头信息失败'
            return False

    # 总逻辑
    def test(self):
        if self.check():
            if self.check_z():
                if self.open_ssh():
                    self.check_position()
        if self.result_failed != '':
            print(u'IP:{0},执行失败,失败原因:{1}'.format(self.ip,self.result_failed))
        else:
            print(u'IP:{0},执行成功通过'.format(self.ip))
        if '\r\n' in self.result:
            self.result = self.result[:-2]
        if '\r\n' in self.result_failed:
            self.result_failed = self.result_failed[:-2]


# 读取Excel信息
class Excel:

    # 初始化
    def __init__(self):
        self.filename = 'devices/devices.xlsx'  # 读取的文件名
        self.ip = []                            # 设备IP
        self.username = []
        self.password = []
        self.port = []                         # 格式化次数


    def read(self):
        # 读取Excel
        try:
            data = openpyxl.load_workbook(self.filename)

            # 读取Sheet1表
            read = data[u'电动镜头位置获取']

            # 获取设备IP与格式化次数
            self.ip = list(filter(None, getColValues(read, 1)))[1:]
            self.username = list(filter(None, getColValues(read, 2)))[1:]
            self.password = list(filter(None, getColValues(read, 3)))[1:]
            self.port = list(filter(None, getColValues(read, 4)))[1:]
            return True
        except Exception as e:
            print(e)
            return False

    # 执行结果写入
    def write(self, model, result, result_failed, CurZoomPos, CurFocusPos):

        # 背景红色填充
        fill = PatternFill("solid", fgColor="ff0000")

        # 设置字体格式
        font = Font(name=u"微软雅黑", size=13, color=colors.BLACK)
        border = Border(left=Side(style='medium', color=colors.BLACK),
                        right=Side(style='medium', color=colors.BLACK),
                        top=Side(style='medium', color=colors.BLACK),
                        bottom=Side(style='medium', color=colors.BLACK))

        # 打开Excel，设置Sheet页名为Sheet
        data = openpyxl.Workbook()
        write = data['Sheet']

        # 遍历写入执行结果
        lenth = len(self.ip)
        for i in range(lenth):
            write.cell(row=i + 2, column=1).value = self.ip[i]
            write.cell(row=i + 2, column=2).value = model[i]
            write.cell(row=i + 2, column=3).value = result[i]
            if 'Failed' in result[i]:
                write.cell(row=i + 2, column=3).fill = fill
            write.cell(row=i + 2, column=4).value = result_failed[i]
            write.cell(row=i + 2, column=5).value = CurZoomPos[i]
            write.cell(row=i + 2, column=6).value = CurFocusPos[i]

        # 为数据设置格式
        for cols in write['A1:F' + str(lenth + 1)]:
            for j in cols:
                j.font = font
                j.border = border

        font_deviceModel = Font(name=u"微软雅黑", size=14, color=colors.BLACK, bold=True)
        write["A1"].font = font_deviceModel
        write.column_dimensions['A'].width = 15
        write.cell(row=1, column=1).value = u"设备IP"

        write["B1"].font = font_deviceModel
        write.column_dimensions['B'].width = 30
        write.cell(row=1, column=2).value = u"设备型号"

        write["C1"].font = font_deviceModel
        write.column_dimensions['C'].width = 13
        write.cell(row=1, column=3).value = u"测试结果"

        write["D1"].font = font_deviceModel
        write.column_dimensions['D'].width = 30
        write.cell(row=1, column=4).value = u"未通过原因"

        write["E1"].font = font_deviceModel
        write.column_dimensions['E'].width = 18
        write.cell(row=1, column=5).value = u"CurZoomPos"

        write["F1"].font = font_deviceModel
        write.column_dimensions['F'].width = 18
        write.cell(row=1, column=6).value = u"CurFocusPos"

        write.column_dimensions["A"].auto_size = True
        write.column_dimensions["B"].auto_size = True
        write.column_dimensions["C"].auto_size = True
        write.column_dimensions["D"].auto_size = True
        write.column_dimensions["E"].auto_size = True
        write.column_dimensions["F"].auto_size = True
        report_name = 'devices/zoom_report__%s.xlsx' % datetime.now().strftime('%Y-%m-%d-%H-%M-%S')
        data.save(report_name)
        print(u'测试结果已导出至工具目录下的{0}'.format(report_name))


# 多线程执行
class MyThread(threading.Thread, Device):

    def __init__(self, ip, username, password, port):
        threading.Thread.__init__(self)
        Device.__init__(self, ip, username, password, port)

    def run(self):
        self.test()


# 从读取到生成报告总流程
def format_stability():
    try:
        print(u'开始执行')
        result_r = []
        excel = Excel()
        if excel.read():
            nums = len(excel.ip)
            li = []
            for i in range(nums):
                t = MyThread(excel.ip[i], excel.username[i], excel.password[i], excel.port[i])
                li.append(t)
                t.start()
            for t in li:
                t.join()
            model = []
            result = []
            result_failed = []
            CurZoomPos = []
            CurFocusPos = []
            for i in range(nums):
                model.append(li[i].model)
                result.append(li[i].result)
                result_failed.append(li[i].result_failed)
                CurZoomPos.append(li[i].CurZoomPos)
                CurFocusPos.append(li[i].CurFocusPos)
            excel.write(model, result, result_failed, CurZoomPos, CurFocusPos)
            for i in li:
                result_r.append(i.result_r)
        print(u"测试结束!!")
        return result_r
    except Exception as e:
        print(e)

if __name__ == '__main__':
    format_stability()
    # a = Device('10.65.74.201', 'admin', 'abcd1234', '22')
    # a.open_ssh()