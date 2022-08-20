# -*- coding: utf-8 -*-
import threading
import openpyxl
from openpyxl.styles import Font, colors, PatternFill, Border,Side
import requests                           # 请求相关的库
from requests.auth import HTTPDigestAuth  # 摘要认证
import time
from lxml import etree
from datetime import datetime
from public_function import getColValues


class Device(object):

    """
    ip:设备ip
    port:设备HTTP端口号
    username：设备登录账号
    password:设备密码
    cout:格式化次数
    """

    # 初始化用户输入的ip、HTTP端口号、用户名、密码等
    def __init__(self, ip, port, username, password, count):
        self.ip = ip                      # 设备IP地址
        self.port = port                  # 设备HTTP端口号
        self.username = username          # 用户名
        self.password = password          # 密码
        self.count = int(count)           #格式化次数
        self.result = ''                  # 执行结果
        self.result_failed = ''           # 失败原因
        self.storage = []                 # 设备存储介质个数
        self.storage_int = []             # 分割nas和sd卡
        self.storage_hdd = []             # 设备SD卡数量
        self.storage_nas = []             # 设备nas数量
        self.result_r = u'执行通过'        # 打印的执行结果

    # 登录设备
    def check(self):
        n = 0
        while True:
            try:
                n += 1
                get = requests.get(u'http://{0}:{1}{2}'.format(self.ip, self.port, u'/ISAPI/System/deviceInfo'),
                                   auth=HTTPDigestAuth(self.username, self.password), timeout=5)
                get.raise_for_status()
                return True
            except Exception as e:
                if n >= 2:
                    self.result += u'Failed'
                    self.result_r = u"{0}：登录失败，请检查是否在线".format(self.ip)
                    self.result_failed += u'设备登录失败，请检查'
                    return False

    # 获取设备存储介质信息
    def storage_information(self):
        nas_get = requests.get(u'http://{0}:{1}{2}'.format(self.ip, self.port, '/ISAPI/ContentMgmt/Storage'),
                           auth=HTTPDigestAuth(self.username, self.password), timeout=5)
        nas_get.raise_for_status()
        root = etree.XML(nas_get.text.encode('utf-8'))
        check_hdd = root.xpath("//*[local-name()='hdd']/*[local-name()='id']")
        check_nas = root.xpath("//*[local-name()='nas']/*[local-name()='id']")

        if check_hdd:
            for i in check_hdd:
                self.storage.append(i.text)
        if check_nas:
            for i in check_nas:
                self.storage.append(i.text)
        for i in self.storage:
            self.storage_int.append(int(i))
        for i in self.storage_int:
            if i < 9 :
                self.storage_hdd.append(str(i))
            else:
                self.storage_nas.append(str(i))

    # 格式化SD卡
    def format_hdd(self, id):
        n = 0
        while n < self.count:
            try:
                n += 1
                put = requests.put(
                    u'http://%s:%s' % (self.ip, self.port) + "/ISAPI/ContentMgmt/Storage/hdd/%s/format" % id,
                    auth=HTTPDigestAuth(self.username, self.password), timeout=600)
                put.raise_for_status()
                time.sleep(3)
                get = requests.get(u'http://%s:%s' % (self.ip, self.port) + "/ISAPI/ContentMgmt/Storage/hdd/%s" % id,
                                   auth=HTTPDigestAuth(self.username, self.password), timeout=5)
                get.raise_for_status()
                root = etree.XML(get.text.encode('utf-8'))
                check = root.xpath("//*[local-name()='status']")[0].text
                if check != u'ok':
                    self.result += u'No.{0}-SD卡-Failed\r\n'.format(id)
                    self.result_r = u"{0}：第{1}次SD卡格式化成功，但状态异常".format(self.ip,n)
                    now_time = datetime.now().strftime("%Y-%m-%dT%H:%M:%S")
                    self.result_failed = u'NO.{0}-SD卡第{1}次SD卡格式化成功，但状态异常\r\n失败时间：{2}'.format(id, n, now_time)
                    return False
            except Exception as e:
                self.result += u'No.{0}-SD卡-Failed\r\n'.format(id)
                self.result_r = u"{0}：第{1}次SD卡无法格式化".format(self.ip, n)
                now_time = datetime.now().strftime("%Y-%m-%dT%H:%M:%S")
                self.result_failed += u'NO.{0}-SD卡第{1}次无法格式化，直接结束测试\r\n失败时间：{2}\r\n'.format(id, n, now_time)
                return False
        self.result += u'Pass'
        return True

    # 格式化NAS
    def format_nas(self, id):
        n = 0
        while n < self.count:
            try:
                n += 1
                put = requests.put(
                    u'http://%s:%s' % (self.ip, self.port) + "/ISAPI/ContentMgmt/Storage/nas/%s/format" % id,
                    auth=HTTPDigestAuth(self.username, self.password), timeout=600)
                put.raise_for_status()
                time.sleep(3)
                get = requests.get(u'http://%s:%s' % (self.ip, self.port) + "/ISAPI/ContentMgmt/Storage/nas/",
                                   auth=HTTPDigestAuth(self.username, self.password), timeout=5)
                get.raise_for_status()
                root = etree.XML(get.text.encode('utf-8'))
                check = root.xpath("//*[text()='{0}']/parent::*/*[local-name()='status']".format(id))[0].text
                if check != u'ok':
                    self.result += u'No.{0}-NAS-Failed\r\n'.format(id)
                    self.result_r = u"{0}：第{1}次NAS格式化成功，但状态异常".format(self.ip, n)
                    now_time = datetime.now().strftime("%Y-%m-%dT%H:%M:%S")
                    self.result_failed += u'NO.{0}-NAS第{1}次格式化成功，但状态异常\r\n失败时间：{2}\r\n'.format(id, n, now_time)
                    return False
            except Exception as e:
                self.result += u'No.{0}-NAS-Failed\r\n'.format(id)
                self.result_r = u"{0}：第{1}次NAS无法格式化".format(self.ip, n)
                now_time = datetime.now().strftime("%Y-%m-%dT%H:%M:%S")
                self.result_failed += u'NO.{0}-NAS第{1}次无法格式化，直接结束测试\r\n失败时间：{2}\r\n'.format(id, n,now_time)
                return False
        self.result += u'No.{0}-NAS-Pass\r\n'.format(id)
        return True

    # 格式化流程逻辑
    def format_storage(self):
        if self.storage_hdd:
            for i in self.storage_hdd:
                self.format_hdd(i)
        elif self.storage_nas:
            for i in self.storage_nas:
                self.format_nas(i)

    # 总逻辑
    def test(self):
        print(u'IP:{0},开始执行'.format(self.ip))
        if self.check():
            self.storage_information()
            if self.storage:
                self.format_storage()
            else:
                self.result += u'Failed'
                self.result_failed += u'该设备未检查到存储介质，请检查'
                self.result_r = u"{0}：未检查到存储介质，请检查".format(self.ip)
        if self.result_failed != '':
            print(u'IP:{0},执行失败,失败原因:{1}'.format(self.ip,self.result_failed))
        else:
            print(u'IP:{0},执行成功通过'.format(self.ip))
        if '\r\n' in self.result:
            self.result = self.result[:-2]
        if '\r\n' in self.result_failed:
            self.result_failed = self.result_failed[:-2]


# 读取Excel信息
class Excel:

    # 初始化
    def __init__(self):
        self.filename = 'devices/devices.xlsx'  # 读取的文件名
        self.ip = []                            # 设备IP
        self.count = []                         # 格式化次数


    def read(self):
        # 读取Excel
        try:
            data = openpyxl.load_workbook(self.filename)

            # 读取Sheet1表
            read = data[u'大面积格式化']

            # 获取设备IP与格式化次数
            self.ip = list(filter(None, getColValues(read, 1)))[1:]
            self.username = list(filter(None, getColValues(read, 2)))[1:]
            self.password = list(filter(None, getColValues(read, 3)))[1:]
            self.count = list(filter(None, getColValues(read, 4)))[1:]
            return True
        except Exception as e:
            print(e)
            return False

    # 执行结果写入
    def write(self, result, result_failed):

        # 背景红色填充
        fill = PatternFill("solid", fgColor="ff0000")

        # 设置字体格式
        font = Font(name=u"微软雅黑", size=13, color=colors.BLACK)
        border = Border(left=Side(style='medium', color=colors.BLACK),
                        right=Side(style='medium', color=colors.BLACK),
                        top=Side(style='medium', color=colors.BLACK),
                        bottom=Side(style='medium', color=colors.BLACK))

        # 打开Excel，设置Sheet页名为Sheet
        data = openpyxl.Workbook()
        write = data['Sheet']

        # 遍历写入执行结果
        lenth = len(self.ip)
        for i in range(lenth):
            write.cell(row=i + 2, column=1).value = self.ip[i]
            write.cell(row=i + 2, column=2).value = result[i]
            if 'Failed' in result[i]:
                write.cell(row=i + 2, column=2).fill = fill
            write.cell(row=i + 2, column=3).value = result_failed[i]

        # 为数据设置格式
        for cols in write['A1:C' + str(lenth + 1)]:
            for j in cols:
                j.font = font
                j.border = border

        font_deviceModel = Font(name=u"微软雅黑", size=14, color=colors.BLACK, bold=True)
        write["A1"].font = font_deviceModel
        write.column_dimensions['A'].width = 15
        write.cell(row=1, column=1).value = u"设备IP"

        write["B1"].font = font_deviceModel
        write.column_dimensions['B'].width = 18
        write.cell(row=1, column=2).value = u"测试结果"

        write["C1"].font = font_deviceModel
        write.column_dimensions['C'].width = 40
        write.cell(row=1, column=3).value = u"未通过原因"

        write.column_dimensions["A"].auto_size = True
        write.column_dimensions["B"].auto_size = True
        write.column_dimensions["C"].auto_size = True
        report_name = 'devices/report__%s.xlsx' % datetime.now().strftime('%Y-%m-%d-%H-%M-%S')
        data.save(report_name)
        print(u'测试结果已导出至工具目录下的{0}'.format(report_name))


# 多线程执行
class MyThread(threading.Thread, Device):

    def __init__(self, ip, port, username, password, count):
        threading.Thread.__init__(self)
        Device.__init__(self, ip, port, username, password, count)

    def run(self):
        self.test()


# 从读取到生成报告总流程
def format_stability():
    try:
        print(u'开始执行')
        result_r = []
        excel = Excel()
        if excel.read():
            nums = len(excel.ip)
            li = []
            for i in range(nums):
                t = MyThread(excel.ip[i], '80', excel.username[i], excel.password[i], excel.count[i])
                li.append(t)
                t.start()
            for t in li:
                t.join()
            result = []
            result_failed = []
            for i in range(nums):
                result.append(li[i].result)
                result_failed.append(li[i].result_failed)
            excel.write(result, result_failed)
            for i in li:
                result_r.append(i.result_r)
        print(u"测试结束!!")
        return result_r
    except Exception as e:
        print(e)

if __name__ == '__main__':
    format_stability()